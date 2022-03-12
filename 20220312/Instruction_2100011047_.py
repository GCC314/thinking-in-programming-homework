import openpyxl
import copy
from openpyxl.chart import BarChart, Series, Reference

#Read datas
wBook = openpyxl.Workbook()
fData = open("asmData.txt","r")
rawData = fData.read()
linesData = rawData.splitlines(False)

#Filter and split datas
codes = []
for line in linesData:
    code = line.split()
    if(len(code) <= 3):
        continue
    codes.append(code)
#idk which code to be ignored

'''
code structure:
    linen   addr    ctln    cname   args...
'''

#Phase 1
wBook.create_sheet(index=0,title="instructionCount")
sheet = wBook.worksheets[0]
countMap = {}
for code in codes:
    if(code[3] in countMap):
        countMap[code[3]] += 1
    else:
        countMap[code[3]] = 1
needTitle = True # enable needTitle to place a title before data
idx = 1
if(needTitle):
    sheet.cell(idx,1).value , sheet.cell(idx,2).value = "Instruction" , "Count"
    idx += 1
for ins,cnt in countMap.items():
    sheet.cell(idx,1).value , sheet.cell(idx,2).value = ins , cnt
    idx += 1

#Phase 2

'''
In order to simplify the problem,we make the extra rules below:
    1)  Arg `return;` after instruction `jxx` will be ignored
    2)  We regard `jmp patch_lazy_pointers(mach_header	const*,	patch_t*,	unsigned	long)`
        as an instruction with **one** argument.
    3)  Instructions with the same code name and different argument numbers
        were regarded as **different** kinds of instructions in **Phase 2 & 3**
'''
def FilterCount(code):
    cnt = len(code) - 4
    if(code[3].startswith("j")):
        if(cnt > 1 and code[5] == "return;"):
            cnt -= 1
    if(code[3] == "jmp" and cnt > 4):
        cnt = 1
    return cnt

wBook.create_sheet(index=1,title="instructionType")
sheet = wBook.worksheets[1]
argMap = {}
repMap = {}
for code in codes:
    cname = code[3]
    argc = FilterCount(code)
    if(argc >= 3):
        argc = 3 # represents k-args
    if(not argc in argMap):
        argMap[argc] = set()
    argMap[argc].add(cname)
col = 1
for argc,cnames in argMap.items():
    row = 1
    if(argc > 2):
        sheet.cell(row,col).value = "3+ Arguments"
    else:
        sheet.cell(row,col).value = str(argc) + " Arguments"
    row += 1
    for cname in cnames:
        if(cname in repMap):
            repMap[cname] += 1
        else:
            repMap[cname] = 1
        sheet.cell(row,col).value = cname
        row += 1
    col += 1
for k,v in repMap.items():
    if(v > 1):
        print(k,v) # for debug

#Phase 3

def tName(x):
    if(x >= 3):
        return "3+ Arguments"
    else:
        return str(x) + " Arguments"

wBook.create_sheet(index=2,title="Summary")
sheet = wBook.worksheets[2]
needTitle = True
idx = 1
sumMap = {}
for code in codes:
    ct = (code[3],FilterCount(code))
    if(ct in sumMap):
        sumMap[ct] += 1
    else:
        sumMap[ct] = 1
sumList = list(sumMap.items())
sumList.sort(key=lambda x:x[1],reverse=True)
if(needTitle):
    sheet.cell(idx,1).value , sheet.cell(idx,2).value , sheet.cell(idx,3).value = "Instruction" , "Amount" , "Type"
    idx += 1
for ct in sumList:
    sheet.cell(idx,1).value , sheet.cell(idx,2).value , sheet.cell(idx,3).value = ct[0][0],ct[1],tName(ct[0][1])
    idx += 1

#Phase 4
wBook.create_sheet(index=3,title="Chart")
sheet = wBook.worksheets[3]
cntType = [0] * 4
for code in codes:
    cntType[min(3,FilterCount(code))] += 1
rows = []
needTitle = True
if(needTitle):
    rows.append(("Args","Count"))
for i in range(0,4):
    rows.append((tName(i),cntType[i]))
for row in rows:
    sheet.append(row)
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Bar Chart of numbers of instructions"
chart.y_axis.title = 'Numbers'
chart.x_axis.title = 'Arguments'
data = Reference(sheet, min_col=2, min_row=1, max_row=5, max_col=2)
cats = Reference(sheet, min_col=1, min_row=2, max_row=5)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
sheet.add_chart(chart, "A10")

#End
wBook.save("Instruction_2100011047_.xlsx")